import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

excel = pd.read_excel('supermarket_sales.xlsx')

tablap = excel.pivot_table(index = 'Gender', columns='Product line', values='Total', aggfunc='sum').round(0)
tablap.to_excel('Proyecto_8_reporte.xlsx', startrow=4, sheet_name='Report')

wb = load_workbook('Proyecto_8_reporte.xlsx')
pestaña = wb['Report']

columna_minima = wb.active.min_column
columna_maxima = wb.active.max_column
fila_minima = wb.active.min_row
fila_maxima = wb.active.max_row

grafico = BarChart()

data = Reference(pestaña, min_col=columna_minima+1, max_col=columna_maxima, min_row=fila_minima, max_row=fila_maxima)
categorias = Reference(pestaña, min_col=columna_minima, max_col=columna_maxima, min_row=fila_minima+1, max_row=fila_maxima)

grafico.add_data(data, titles_from_data=True)
grafico.set_categories(categorias)

pestaña.add_chart(grafico,'B12')
grafico.title='Ventas'
grafico.style = 2

wb.save('Proyecto_8_reporte.xlsx')
